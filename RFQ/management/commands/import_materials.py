import os
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from RFQ.models import Material

class Command(BaseCommand):
    help = 'Importa resinas industriales desde un archivo Excel (.xlsx) de forma segura'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Ruta al archivo Excel')

    def handle(self, *args, **options):
        excel_path = options['excel_file']
        
        if not os.path.exists(excel_path):
            self.stdout.write(self.style.ERROR(f"El archivo no existe en la ruta: {excel_path}"))
            return

        self.stdout.write(self.style.SUCCESS("Leyendo libro de Excel..."))
        
        try:
            # data_only=True asegura que leamos los valores finales y no las fórmulas de Excel
            wb = load_workbook(excel_path, data_only=True)
            sheet = wb.active  # Toma la primera pestaña activa del Excel
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error al abrir el archivo Excel: {e}"))
            return

        self.stdout.write(self.style.SUCCESS(f"Procesando pestaña: '{sheet.title}'. Mapeando columnas..."))
        
        # Leemos la primera fila para mapear qué columna tiene cada dato
        header = [cell.value for cell in sheet[1]]
        
        try:
            col_code = header.index('material_code')
            col_family = header.index('family')
            col_name = header.index('commercial_name')
            col_supplier = header.index('supplier')
            col_color = header.index('color')
            col_density = header.index('density')
        except ValueError as ve:
            self.stdout.write(self.style.ERROR(
                f"Error de cabecera: Asegúrate de que la primera fila del Excel tenga exactamente los nombres: "
                f"'material_code', 'family', 'commercial_name', 'supplier', 'color', 'density'. Detalles: {ve}"
            ))
            return

        count_created = 0
        count_skipped = 0

        # Iteramos desde la fila 2 para saltarnos los encabezados
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Si toda la fila está vacía, nos la saltamos
            if not any(row):
                continue
                
            try:
                code = str(row[col_code]).strip() if row[col_code] is not None else ""
                family = str(row[col_family]).strip() if row[col_family] is not None else ""
                name = str(row[col_name]).strip() if row[col_name] is not None else ""
                supplier = str(row[col_supplier]).strip() if row[col_supplier] is not None else ""
                color = str(row[col_color]).strip() if row[col_color] is not None else ""
                
                # Validación segura de la densidad matemática
                try:
                    density = float(row[col_density]) if row[col_density] is not None else 1.05
                except (ValueError, TypeError):
                    density = 1.05

                if not code or code.upper() == 'NONE':
                    continue

                # Guardamos o actualizamos en la base de datos de Railway
                obj, created = Material.objects.update_or_create(
                    material_code=code,
                    defaults={
                        'family': family,
                        'commercial_name': name,
                        'supplier': supplier,
                        'color': color,
                        'density': density
                    }
                )
                
                if created:
                    count_created += 1
                else:
                    count_skipped += 1

            except Exception as row_err:
                continue

        self.stdout.write(self.style.SUCCESS(
            f"¡Migración desde Excel exitosa! Registros nuevos: {count_created} | Actualizados/Omitidos: {count_skipped}"
        ))